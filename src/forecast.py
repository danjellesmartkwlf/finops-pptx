"""Forecast ingestion from Excel workbooks.

This module handles loading forecast data from Excel files as specified
in the project config.yaml.  The forecast spreadsheet is expected to have
row labels in column A and monthly values across subsequent columns, with
column headers containing parseable month/year information (e.g.
"Jan 2026", "January 2026", "2026-01-01").
"""

from __future__ import annotations

import calendar
import datetime
import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_column_date(header_value: Any) -> tuple[int, int] | None:
    """Attempt to extract (month, year) from a column header value.

    Supports:
        - datetime / date objects (as openpyxl may return)
        - Strings like "Jan 2026", "January 2026", "2026-01-01", "1/2026"

    Args:
        header_value: The raw cell value from the header row.

    Returns:
        A (month, year) tuple, or ``None`` if parsing fails.
    """
    if header_value is None:
        return None

    # datetime objects returned directly by openpyxl
    if isinstance(header_value, (datetime.datetime, datetime.date)):
        return (header_value.month, header_value.year)

    text = str(header_value).strip()
    if not text:
        return None

    # Try ISO-style: "2026-01-01" or "2026-01"
    for fmt in ("%Y-%m-%d", "%Y-%m"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return (dt.month, dt.year)
        except ValueError:
            continue

    # Try "January 2026" / "Jan 2026"
    for fmt in ("%B %Y", "%b %Y"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return (dt.month, dt.year)
        except ValueError:
            continue

    # Try "1/2026" or "01/2026"
    for fmt in ("%m/%Y",):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return (dt.month, dt.year)
        except ValueError:
            continue

    return None


def _resolve_month_year(month: str, year: int) -> tuple[int, int]:
    """Convert a human-readable month string and year into (month_number, year).

    Args:
        month: Month name (full or abbreviated), e.g. "January" or "Jan".
        year: Four-digit year.

    Returns:
        A (month_number, year) tuple.

    Raises:
        ValueError: If the month string cannot be parsed.
    """
    month_lower = month.strip().lower()

    # Full month names
    for i, name in enumerate(calendar.month_name):
        if name and name.lower() == month_lower:
            return (i, year)

    # Abbreviated month names
    for i, name in enumerate(calendar.month_abbr):
        if name and name.lower() == month_lower:
            return (i, year)

    raise ValueError(
        f"Cannot parse month '{month}'. Expected a full or abbreviated "
        f"English month name (e.g. 'January' or 'Jan')."
    )


def _find_target_column(
    sheet: Worksheet,
    target_month: int,
    target_year: int,
) -> int:
    """Locate the column index whose header matches the requested month/year.

    Scans row 1 of *sheet* for a header that parses to the given month and
    year.

    Args:
        sheet: The openpyxl worksheet to scan.
        target_month: Numeric month (1-12).
        target_year: Four-digit year.

    Returns:
        The 1-based column index.

    Raises:
        ValueError: If no matching column is found. The error message lists
            all parseable column headers for debugging.
    """
    available_columns: list[str] = []

    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col_idx).value
        parsed = _parse_column_date(header)
        if parsed is not None:
            m, y = parsed
            if m == target_month and y == target_year:
                return col_idx
            month_name = calendar.month_abbr[m]
            available_columns.append(f"{month_name} {y}")

    available = ", ".join(available_columns) if available_columns else "(none found)"
    raise ValueError(
        f"No column found for month={target_month}, year={target_year}. "
        f"Available date columns: {available}"
    )


def _build_row_label_index(sheet: Worksheet) -> dict[str, int]:
    """Build a mapping of column-A labels to their row numbers.

    Args:
        sheet: The openpyxl worksheet.

    Returns:
        A dict mapping stripped label text to 1-based row number.
    """
    index: dict[str, int] = {}
    for row_idx in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=1).value
        if value is not None:
            index[str(value).strip()] = row_idx
    return index


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _month_column_key(month: str, year: int) -> str:
    """Convert a full month name and year to the column header format used in forecast files.

    Args:
        month: Full or abbreviated month name (e.g. "January" or "Jan").
        year: Four-digit year (e.g. 2026).

    Returns:
        Column header string in ``Mon-YYYY`` format (e.g. "Jan-2026").

    Raises:
        ValueError: If the month string cannot be parsed.
    """
    month_num, yr = _resolve_month_year(month, year)
    return f"{calendar.month_abbr[month_num]}-{yr}"


def _sum_column(
    file_path: Path,
    sheet_name: str,
    column_key: str,
    label: str,
) -> float | None:
    """Read an Excel forecast file and sum all numeric values in the target month column.

    Rows whose first column is NaN are skipped (e.g. total rows).

    Args:
        file_path: Path to the Excel workbook.
        sheet_name: Name of the worksheet to read.
        column_key: The column header to sum (e.g. "Jan-2026").
        label: Human-readable label for log messages (e.g. "COGS" or "OpEx").

    Returns:
        The sum as a float, or ``None`` if the file is missing or the column
        does not exist.
    """
    if not file_path.exists():
        logger.warning("Forecast file not found for %s: %s", label, file_path)
        return None

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        logger.warning(
            "Failed to read sheet '%s' from %s forecast file: %s",
            sheet_name,
            label,
            file_path,
            exc_info=True,
        )
        return None

    if column_key not in df.columns:
        logger.warning(
            "Month column '%s' not found in %s forecast file (%s). "
            "Available columns: %s",
            column_key,
            label,
            file_path,
            list(df.columns),
        )
        return None

    # Drop rows where the first column (category/app label) is NaN
    first_col = df.columns[0]
    df_filtered = df[df[first_col].notna()]

    return float(df_filtered[column_key].sum())


def load_forecasts_from_data_files(
    config: dict[str, Any],
    month: str,
    year: int,
) -> dict[str, float | None]:
    """Load COGS and OpEx forecasts from separate Excel files defined in ``data_files``.

    Reads the COGS and OpEx forecast workbooks specified under
    ``config["data_files"]["forecasts"]``, sums all numeric values in the
    target month column for each file, and computes the total.

    Args:
        config: The full application config dict (parsed from config.yaml).
            Must contain a ``data_files.forecasts`` section with ``cogs``
            and ``opex`` sub-keys, each having ``file_path`` and
            ``sheet_name``.
        month: Month name (full or abbreviated), e.g. "January" or "Jan".
        year: Four-digit year, e.g. 2026.

    Returns:
        A dict with keys ``"total_spend"``, ``"cogs_spend"``, and
        ``"opex_spend"``, each mapped to a float or ``None`` if the
        corresponding file/column could not be read.
    """
    column_key = _month_column_key(month, year)
    forecasts_cfg = config["data_files"]["forecasts"]

    # --- COGS ---
    cogs_cfg = forecasts_cfg["cogs"]
    cogs_spend = _sum_column(
        file_path=Path(cogs_cfg["file_path"]),
        sheet_name=cogs_cfg["sheet_name"],
        column_key=column_key,
        label="COGS",
    )

    # --- OpEx ---
    opex_cfg = forecasts_cfg["opex"]
    opex_spend = _sum_column(
        file_path=Path(opex_cfg["file_path"]),
        sheet_name=opex_cfg["sheet_name"],
        column_key=column_key,
        label="OpEx",
    )

    # --- Total ---
    if cogs_spend is not None and opex_spend is not None:
        total_spend: float | None = cogs_spend + opex_spend
    else:
        total_spend = None

    return {
        "total_spend": total_spend,
        "cogs_spend": cogs_spend,
        "opex_spend": opex_spend,
    }


def load_forecast(
    config: dict[str, Any],
    month: str,
    year: int,
) -> dict[str, float | None]:
    """Load forecast values for a given month from an Excel workbook.

    Opens the workbook and sheet described in ``config["forecast"]``, locates
    the column matching *month*/*year*, and reads the forecast value for each
    mapping key defined in ``config["forecast"]["mapping"]``.

    Args:
        config: The full application config dict (parsed from config.yaml).
        month: Month name (full or abbreviated), e.g. "January" or "Jan".
        year: Four-digit year, e.g. 2026.

    Returns:
        A dict mapping each forecast key (e.g. ``"total_spend"``) to its
        forecast value as a float, or ``None`` when the cell is empty.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        ValueError: If the sheet name is invalid, the month column cannot be
            found, or a configured row label is missing from the spreadsheet.
    """
    # --- New path: separate COGS / OpEx files under data_files ---
    if "data_files" in config and "forecasts" in config.get("data_files", {}):
        logger.debug(
            "Using data_files forecast path (separate COGS/OpEx files)."
        )
        return load_forecasts_from_data_files(config, month, year)

    # --- Legacy path: single forecast file with row-label mapping ---
    forecast_cfg = config["forecast"]
    file_path = Path(forecast_cfg["file_path"])

    if not file_path.exists():
        raise FileNotFoundError(
            f"Forecast Excel file not found: {file_path.resolve()}"
        )

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)

    sheet_name: str = forecast_cfg["sheet_name"]
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available sheets: {available}"
        )

    sheet = wb[sheet_name]

    # Resolve the target month/year
    target_month, target_year = _resolve_month_year(month, year)

    # Find the column for this month/year
    col_idx = _find_target_column(sheet, target_month, target_year)

    # Build a row-label lookup from column A
    row_labels = _build_row_label_index(sheet)

    mapping: dict[str, dict[str, Any]] = forecast_cfg.get("mapping", {})
    results: dict[str, float | None] = {}
    errors: list[str] = []

    for key, key_cfg in mapping.items():
        row_label: str = key_cfg["row_label"]
        if row_label not in row_labels:
            errors.append(
                f"Row label '{row_label}' (for mapping key '{key}') "
                f"not found in column A of sheet '{sheet_name}'."
            )
            continue

        row_idx = row_labels[row_label]
        cell_value = sheet.cell(row=row_idx, column=col_idx).value

        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            results[key] = None
        else:
            try:
                results[key] = float(cell_value)
            except (TypeError, ValueError):
                results[key] = None

    wb.close()

    if errors:
        raise ValueError("\n".join(errors))

    return results


def load_app_forecasts(
    config: dict[str, Any],
    month: str,
    year: int,
) -> dict[str, float]:
    """Load per-app COGS forecast values for a given month.

    Reads the COGS forecast workbook (which has one row per ``awn_app``)
    and returns the forecast value for each app in the target month column.

    Args:
        config: The full application config dict (parsed from config.yaml).
        month: Month name (full or abbreviated), e.g. "January".
        year: Four-digit year, e.g. 2026.

    Returns:
        A dict mapping each ``awn_app`` name to its forecast spend as a float.
        Apps with missing or non-numeric values are excluded.
    """
    column_key = _month_column_key(month, year)
    forecasts_cfg = config.get("data_files", {}).get("forecasts", {})
    cogs_cfg = forecasts_cfg.get("cogs", {})

    file_path = Path(cogs_cfg.get("file_path", ""))
    sheet_name = cogs_cfg.get("sheet_name", "Sheet1")

    if not file_path.exists():
        logger.warning("COGS forecast file not found: %s", file_path)
        return {}

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception:
        logger.warning("Failed to read COGS forecast file: %s", file_path, exc_info=True)
        return {}

    # Identify the app column (first column) and month column
    app_col = df.columns[0]  # "awn_app"

    if column_key not in df.columns:
        logger.warning(
            "Month column '%s' not found in COGS forecast. Available: %s",
            column_key,
            list(df.columns),
        )
        return {}

    result: dict[str, float] = {}
    for _, row in df.iterrows():
        app_name = row[app_col]
        if pd.isna(app_name):
            continue
        app_name = str(app_name).strip()
        val = row[column_key]
        if pd.notna(val):
            try:
                result[app_name] = float(val)
            except (TypeError, ValueError):
                continue

    logger.info("Loaded per-app forecasts for %d apps from %s", len(result), file_path)
    return result


def load_forecast_history(
    config: dict[str, Any],
    month: str,
    year: int,
    num_months: int = 6,
) -> dict[str, list[dict[str, Any]]]:
    """Load *num_months* of forecast values for COGS, OpEx, and Total.

    Iterates backwards from the given month/year, calling the existing
    ``_sum_column()`` helper for each month and each forecast file.  Months
    outside the Excel forecast range naturally return ``None``.

    Args:
        config: The full application config dict (parsed from config.yaml).
        month: Full month name (e.g. "January").
        year: Four-digit year (e.g. 2026).
        num_months: How many months to include (default 6).

    Returns:
        A dict keyed by forecast mapping key (``"cogs_spend"``,
        ``"opex_spend"``, ``"total_spend"``).  Each value is a chronological
        list of dicts with ``month_label`` (str, e.g. "Jan 2026") and
        ``forecast`` (float | None).
    """
    # Walk back months chronologically
    target_month, target_year = _resolve_month_year(month, year)
    month_list: list[tuple[int, int]] = []
    m, y = target_month, target_year
    for _ in range(num_months):
        month_list.append((m, y))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    month_list.reverse()

    forecasts_cfg = config.get("data_files", {}).get("forecasts", {})
    cogs_cfg = forecasts_cfg.get("cogs", {})
    opex_cfg = forecasts_cfg.get("opex", {})

    cogs_history: list[dict[str, Any]] = []
    opex_history: list[dict[str, Any]] = []
    total_history: list[dict[str, Any]] = []

    for m_num, y_num in month_list:
        month_name = calendar.month_name[m_num]
        month_abbr = calendar.month_abbr[m_num]
        label = f"{month_abbr} {y_num}"
        column_key = f"{month_abbr}-{y_num}"

        # COGS forecast
        cogs_val: float | None = None
        if cogs_cfg.get("file_path"):
            cogs_val = _sum_column(
                file_path=Path(cogs_cfg["file_path"]),
                sheet_name=cogs_cfg.get("sheet_name", "Sheet1"),
                column_key=column_key,
                label="COGS",
            )
        cogs_history.append({"month_label": label, "forecast": cogs_val})

        # OpEx forecast
        opex_val: float | None = None
        if opex_cfg.get("file_path"):
            opex_val = _sum_column(
                file_path=Path(opex_cfg["file_path"]),
                sheet_name=opex_cfg.get("sheet_name", "Sheet1"),
                column_key=column_key,
                label="OpEx",
            )
        opex_history.append({"month_label": label, "forecast": opex_val})

        # Total = COGS + OpEx when both exist
        total_val: float | None = None
        if cogs_val is not None and opex_val is not None:
            total_val = cogs_val + opex_val
        total_history.append({"month_label": label, "forecast": total_val})

    return {
        "cogs_spend": cogs_history,
        "opex_spend": opex_history,
        "total_spend": total_history,
    }


def validate_forecast_schema(config: dict[str, Any]) -> list[str]:
    """Validate that all expected row labels exist in the forecast spreadsheet.

    Opens the workbook and checks column A of the configured sheet for every
    ``row_label`` defined in ``config["forecast"]["mapping"]``.

    Args:
        config: The full application config dict (parsed from config.yaml).

    Returns:
        A list of error message strings (empty if validation passes).

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        ValueError: If validation errors are found (raised after collecting
            all errors, with messages joined by newlines).  Also raised if
            the configured sheet does not exist in the workbook.
    """
    forecast_cfg = config["forecast"]
    file_path = Path(forecast_cfg["file_path"])

    if not file_path.exists():
        raise FileNotFoundError(
            f"Forecast Excel file not found: {file_path.resolve()}"
        )

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)

    sheet_name: str = forecast_cfg["sheet_name"]
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available sheets: {available}"
        )

    sheet = wb[sheet_name]
    row_labels = _build_row_label_index(sheet)

    mapping: dict[str, dict[str, Any]] = forecast_cfg.get("mapping", {})
    errors: list[str] = []

    for key, key_cfg in mapping.items():
        row_label: str = key_cfg["row_label"]
        if row_label not in row_labels:
            errors.append(
                f"Missing row label '{row_label}' (forecast key '{key}') "
                f"in sheet '{sheet_name}'."
            )

    wb.close()

    if errors:
        raise ValueError("\n".join(errors))

    return errors
